# -*- coding: utf-8 -*-
"""
@Time : 2024/6/4 15:28
@Email : Lvan826199@163.com
@公众号 : 梦无矶的测试开发之路
@File : 15_写入带有超链接的内容.py
"""
__author__ = "梦无矶小仔"

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# 现有的 Excel 文件
excel_file = 'excel_path/write3.xlsx'

# 创建新的 DataFrame 以追加
new_data = {'Name': ['百度', 'CSDN主页'], 'URL': ['https://www.baidu.com', 'https://blog.csdn.net/qq_46158060']}
new_df = pd.DataFrame(new_data)

# 使用 openpyxl 加载现有的工作簿
wb = load_workbook(excel_file)
# ws = wb.active
ws = wb["Sheet1"]

# 找到最后一行
if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
    last_row = 0
else:
    last_row = ws.max_row


# 将新的 DataFrame 追加到现有的 Excel 文件
for r in dataframe_to_rows(new_df, index=False, header=last_row == 0):
    ws.append(r)

# 添加超链接和样式
start_row = last_row + 1 if last_row != 0 else 1
for row in range(start_row, start_row + len(new_df)):
    cell = ws.cell(row=row, column=1)
    url = ws.cell(row=row, column=2).value
    cell.hyperlink = url
    cell.font = Font(color="0000FF", underline="single")

# 保存工作簿
wb.save(excel_file)

