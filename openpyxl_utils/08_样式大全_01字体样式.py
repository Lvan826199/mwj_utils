# -*- coding: utf-8 -*-
"""
@Time : 2024/6/6 17:52
@Email : Lvan826199@163.com
@公众号 : 梦无矶测开实录
@File : 08_样式大全_01字体样式.py
"""
__author__ = "梦无矶小仔"
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active

# 设置字体
font = Font(name='Arial', size=12, bold=True, italic=True, underline='single', color='FF0000')
ws['A1'].font = font
ws['A1'] = "Hello, World!"

wb.save("excelPath/demo08.xlsx")