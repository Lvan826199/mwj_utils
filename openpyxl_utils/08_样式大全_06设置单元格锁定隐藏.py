# -*- coding: utf-8 -*-
"""
@Time : 2024/6/7 11:39
@Email : Lvan826199@163.com
@公众号 : 梦无矶测开实录
@File : 08_样式大全_06设置单元格锁定隐藏.py
"""
__author__ = "梦无矶小仔"

from openpyxl import Workbook
from openpyxl.styles import Protection

wb = Workbook()
ws = wb.active
# 设置单元格保护
protection = Protection(locked=True, hidden=False)
ws['A6'].protection = protection
ws['A6'] = "Protected Cell"

wb.save("excelPath/demo08_6.xlsx")
