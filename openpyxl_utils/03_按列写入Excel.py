# -*- coding: utf-8 -*-
"""
@Time : 2024/6/5 17:55
@Email : Lvan826199@163.com
@公众号 : 梦无矶测开实录
@File : 03_按列写入Excel.py
"""
__author__ = "梦无矶小仔"

import openpyxl
from openpyxl.utils import get_column_letter


def write_data_by_column(data, filename):
    """
    将数据按列写入 Excel 文件。
    :param data: (dict): 包含列标和对应数据的字典，或者包含列数据的嵌套列表。
    :param filename: (str): 要保存的文件名。
    """
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 检查数据类型并写入数据
    if isinstance(data, dict):
        for col, values in data.items():
            for row, value in enumerate(values, start=1):
                ws[f'{col}{row}'] = value

    elif isinstance(data, list):
        for col_idx, col_data in enumerate(data, start=1):
            col_letter = get_column_letter(col_idx)
            for row_idx, value in enumerate(col_data, start=1):
                ws[f'{col_letter}{row_idx}'] = value
    else:
        raise ValueError("数据格式不正确，请提供字典或嵌套列表。")

    # 保存工作簿
    wb.save(filename)


if __name__ == '__main__':
    data_dict = {
        'A': ["Name", "梦无矶", "小仔", "沐默"],
        'B': ["Age", 30, 25, 35],
        'C': ["City", "潮汕", "银川", "台北"]
    }

    data_list = [
        ["Name", "梦无矶", "小仔", "沐默"],
        ["Age", 30, 25, 35],
        ["City", "香港", "璃月", "阿尔及利亚"]
    ]

    # 使用字典数据写入
    write_data_by_column(data_dict, "excelPath/demo3_dict.xlsx")

    # 使用列表数据写入
    write_data_by_column(data_list, "excelPath/demo3_list.xlsx")
