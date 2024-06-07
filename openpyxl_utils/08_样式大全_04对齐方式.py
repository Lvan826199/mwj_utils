# -*- coding: utf-8 -*-
"""
@Time : 2024/6/7 11:26
@Email : Lvan826199@163.com
@公众号 : 梦无矶测开实录
@File : 08_样式大全_04对齐方式.py
"""
__author__ = "梦无矶小仔"

from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active
# 设置对齐方式
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws['A4'].alignment = alignment
ws['A4'] = "Centered Text\n我是第二行"

# 保存工作簿
wb.save("excelPath/demo08_4.xlsx")
