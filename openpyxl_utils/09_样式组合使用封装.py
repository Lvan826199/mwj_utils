# -*- coding: utf-8 -*-
"""
@Time : 2024/6/7 11:46
@Email : Lvan826199@163.com
@公众号 : 梦无矶测开实录
@File : 09_样式组合使用封装.py
"""
__author__ = "梦无矶小仔"

import os
from copy import copy

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Optional


class ExcelWriter:
    def __init__(self, filename):
        self.filename: str = filename
        self.workbook: Workbook = Workbook()
        self.sheet: Worksheet = self.workbook.active

    def save(self):
        self.workbook.save(self.filename)

    def create_excel(self):
        # 新建一个空表格
        self.save()

    def create_excel_with_row(self, row_data: list, has_title=True, styles=None):
        """
        新建一个带有标题的表格,单行写入
        :param row_data:list[]
        :param has_title:
        :param styles:
        :return:
        """
        self.append_row(row_data, has_title=has_title, styles=styles)

    def create_excel_with_rows(self, rows_data: list, has_title=True, styles=None):
        """
        新建一个带有标题的表格,多行写入
        :param rows_data: list[[],[],[]...]
        :param has_title:
        :param styles:
        :return:
        """
        self.append_rows(rows_data, has_title=has_title, styles=styles)

    def verfiy_row_data(self, row_data: list, excel_title):
        """
        校验行数据长度
        :param row_data:
        :param excel_title:
        :return:
        """
        if len(row_data) != len(excel_title):
            raise ValueError(f"{row_data}写入行的数据长度必须和标题行长度一致！")

    def append_row(self, row_data: list, has_title: bool = False, styles=None):
        """
        写入或追加一行数据并设置样式
        :param row_data: list[]
        :param has_title: 是否带有标题
        :param styles:
        :return:
        """
        if has_title:
            for col_num, value in enumerate(row_data, 1):
                cell = self.sheet.cell(row=1, column=col_num, value=value)
                self.write_with_styles(1, col_num, cell, styles)
                self.save()
        else:
            row_num = self.sheet.max_row + 1
            for col_num, value in enumerate(row_data, 1):
                cell = self.sheet.cell(row=row_num, column=col_num, value=value)
                self.write_with_styles(row_num, col_num, cell, styles)
            self.save()

    def append_rows(self, rows_data: list, has_title: bool = False, styles=None):
        """
        写入或追加多行数据并设置样式
        :param row_data: list[[],[],[]...]
        :param has_title: 是否带有标题
        :param styles:
        :return:
        """

        if has_title and isinstance(rows_data[0], list):
            for col_num, value in enumerate(rows_data[0], 1):
                cell = self.sheet.cell(row=1, column=col_num, value=value)
                self.write_with_styles(1, col_num, cell, styles)
                self.save()
            for row_data in rows_data[1:]:
                row_num = self.sheet.max_row + 1
                for col_num, value in enumerate(row_data, 1):
                    cell = self.sheet.cell(row=row_num, column=col_num, value=value)
                    self.write_with_styles(row_num, col_num, cell, styles)
                self.save()
        else:
            for row_data in rows_data:
                row_num = self.sheet.max_row + 1
                for col_num, value in enumerate(row_data, 1):
                    cell = self.sheet.cell(row=row_num, column=col_num, value=value)
                    self.write_with_styles(row_num, col_num, cell, styles)
                self.save()

    def copy_excel_to_sheet(self, source_excel_path, sheet_name):
        """
        复制表数据不包含样式到指定sheet页
        :param source_excel_path:
        :param sheet_name:
        :return:
        """
        # 打开源文件和目标文件
        source_wb = openpyxl.load_workbook(source_excel_path)
        destination_wb = openpyxl.load_workbook(self.filename)

        # 获取源文件中的第一个 sheet
        source_sheet = source_wb.active

        # 在目标文件中创建一个新的 sheet
        new_sheet = destination_wb.create_sheet(title=sheet_name)

        # 遍历源 sheet 中的所有单元格，并将其值复制到新 sheet 中
        for row in source_sheet.iter_rows(values_only=True):
            new_sheet.append(row)

        # 保存目标文件
        destination_wb.save(self.filename)
        run_dir = os.getcwd()
        excel_path = os.path.join(run_dir, self.filename)
        print(f"生成结果表格路径：{excel_path}")

    def copy_excel_to_sheet_with_styles(self, source_excel_path, sheet_name):
        """
        复制表数据包含样式到指定sheet页
        :param source_excel_path:
        :param sheet_name:
        :return:
        """
        # 打开源文件和目标文件
        source_wb = openpyxl.load_workbook(source_excel_path)
        destination_wb = openpyxl.load_workbook(self.filename)

        # 获取源文件中的第一个 sheet
        source_sheet = source_wb.active

        # 在目标文件中创建一个新的 sheet
        new_sheet = destination_wb.create_sheet(title=sheet_name)

        # 复制单元格的值和样式
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                # 复制样式
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # 复制列宽
        for col in source_sheet.column_dimensions:
            new_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width

        # 复制行高
        for row in source_sheet.row_dimensions:
            new_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

        # 保存目标文件
        destination_wb.save(self.filename)
        run_dir = os.getcwd()
        excel_path = os.path.join(run_dir, self.filename)
        print(f"生成结果表格路径：{excel_path}")

    def write_with_styles(self, row_num, col_num, cell, styles=None):
        """
        对行的指定列进行样式设置,不支持多行不同列样式设置
        :param row_num:
        :param col_num:
        :param cell:
        :param styles:
        :return:
        """
        # 设置默认样式
        cell.font = Font(color="000000")
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        # 如果提供了样式，则应用样式
        if styles and col_num in styles:
            print(f"styles:{row_num},{col_num},{cell}")
            style = styles[col_num]
            if 'font_color' in style:
                cell.font = Font(color=style['font_color'])
            if 'bg_color' in style:
                cell.fill = PatternFill(start_color=style['bg_color'], end_color=style['bg_color'], fill_type="solid")
            if 'alignment' in style:
                cell.alignment = Alignment(
                    wrap_text=style['alignment'].get('wrap_text', True),
                    horizontal=style['alignment'].get('horizontal', "center"),
                    vertical=style['alignment'].get('vertical', "center"),
                    indent=style['alignment'].get('indent', 0)
                )
            if 'width' in style:
                self.sheet.column_dimensions[cell.column_letter].width = style['width']
            if 'height' in style:
                self.sheet.row_dimensions[row_num].height = style['height']
            if 'border' in style:
                cell.border = Border(
                    left=Side(border_style=style['border'].get('left', 'thin'), color=style['border'].get('left_color', '000000')),
                    right=Side(border_style=style['border'].get('right', 'thin'), color=style['border'].get('right_color', '000000')),
                    top=Side(border_style=style['border'].get('top', 'thin'), color=style['border'].get('top_color', '000000')),
                    bottom=Side(border_style=style['border'].get('bottom', 'thin'), color=style['border'].get('bottom_color', '000000'))
                )
            if 'font' in style:
                cell.font = Font(
                    name=style['font'].get('name', 'Calibri'),
                    size=style['font'].get('size', 11),
                    bold=style['font'].get('bold', False),
                    italic=style['font'].get('italic', False),
                    underline=style['font'].get('underline', 'none'),
                    strike=style['font'].get('strike', False),
                    color=style['font'].get('color', '000000')
                )
            if 'number_format' in style:
                cell.number_format = style['number_format']


if __name__ == '__main__':
    # 示例使用
    row_data = [["姓名", "年龄", "性别", "职业", "城市", "语言"],
                ["张三", 25, "男", "工程师", "北京", "中文"],
                ["李四", 30, "女", "设计师", "上海", "英文"],
                ["王五", 35, "男", "教师", "广州", "粤语"],
                ["赵六", 40, "女", "医生", "深圳", "国语"],
                ["钱七", 45, "男", "律师", "杭州", "英语"], ["孙八", 50, "女", "护士", "南京", "国语"]]

    styles = {
        1: {
            'font_color': 'FF0000',
            'bg_color': 'FFFF00',
            'alignment': {'horizontal': 'left', 'vertical': 'top', 'wrap_text': True, 'indent': 1},
            'width': 20,
            'height': 30,
            'border': {'left': 'thick', 'left_color': 'FF0000'},
            'font': {'name': 'Arial', 'size': 14, 'bold': True, 'italic': True, 'underline': 'single', 'strike': True, 'color': 'FF0000'}
        },
        3: {
            'font_color': '0000FF',
            'bg_color': '00FF00',
            'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True},
            'width': 15,
            'border': {'bottom': 'dashed', 'bottom_color': '00FF00'}
        },
        5: {
            'font_color': '00FF00',
            'bg_color': 'FF00FF',
            'alignment': {'horizontal': 'right', 'vertical': 'bottom', 'wrap_text': True},
            'height': 25,
            'border': {'top': 'double', 'top_color': '0000FF'},
            'number_format': 'YYYY-MM-DD HH:MM:SS'
        }
    }
    # 多行写入
    excel_writer = ExcelWriter("excelPath/demo09_多行写入.xlsx")
    excel_writer.create_excel_with_rows(row_data, True, styles)

    # 单行写入
    row_data2 = ["姓名", "年龄", "性别", "职业", "城市", "语言"]
    excel_writer2 = ExcelWriter("excelPath/demo09_单行写入.xlsx")
    excel_writer2.create_excel_with_row(row_data2, True, styles)
