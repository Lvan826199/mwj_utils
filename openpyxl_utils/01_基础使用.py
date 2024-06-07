# -*- coding: utf-8 -*-
"""
@Time : 2024/6/5 17:11
@Email : Lvan826199@163.com
@公众号 : 梦无矶测开实录
@File : 01_基础使用.py
"""
__author__ = "梦无矶小仔"

import openpyxl

# 1、创建一个新的工作簿
wb = openpyxl.Workbook()

# 2、获取当前激活的工作表
ws = wb.active

# 3、给工作表命名
ws.title = "Sheet1"

# 4、保存工作簿
wb.save("excelPath/demo1.xlsx")

# 写入单个单元格
ws['A1'] = "Hello"
ws['B1'] = "World"

# 写入多个单元格
data = [
    ["Name", "Age", "City"],
    ["雷神", 30, "稻妻"],
    ["申鹤", 25, "我家"],
    ["凝光", 35, "璃月"]
]

for row in data:
    ws.append(row)

# 保存工作簿
wb.save("excelPath/demo1.xlsx")

# 读取单个单元格的数据
print(ws['A1'].value)

# 读取多个单元格的数据
for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()

# 读取所有数据
for row in ws.iter_rows(values_only=True):
    print(row)

# 合并单元格
ws.merge_cells('A1:C1')
ws['A1'] = "Merged Cell"

# 拆分单元格
ws.unmerge_cells('A1:C1')

# 保存工作簿
wb.save("excelPath/demo1.xlsx")