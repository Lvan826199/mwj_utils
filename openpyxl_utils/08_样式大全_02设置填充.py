# -*- coding: utf-8 -*-
"""
@Time : 2024/6/6 18:51
@Email : Lvan826199@163.com
@公众号 : 梦无矶测开实录
@File : 08_样式大全_02设置填充.py
"""
__author__ = "梦无矶小仔"

from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws = wb.active
# 设置填充
fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ws['A2'].fill = fill
ws['A2'] = "Background Color"

wb.save("excelPath/demo08_2.xlsx")