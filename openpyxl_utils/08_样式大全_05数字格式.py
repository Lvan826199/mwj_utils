# -*- coding: utf-8 -*-
"""
@Time : 2024/6/7 11:32
@Email : Lvan826199@163.com
@公众号 : 梦无矶测开实录
@File : 08_样式大全_05数字格式.py
"""
__author__ = "梦无矶小仔"

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 设置数字格式
ws['A5'].number_format = 'YYYY-MM-DD'
ws['A5'] = '2024-06-05'

wb.save("excelPath/demo08_5.xlsx")