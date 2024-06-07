# -*- coding: utf-8 -*-
"""
@Time : 2024/6/5 18:21
@Email : Lvan826199@163.com
@公众号 : 梦无矶测开实录
@File : 04_追加写入.py
"""
__author__ = "梦无矶小仔"

from openpyxl import load_workbook


def append_rows_to_excel(file_path, rows, sheet_name=None):
    """
    追加行数据到现有的Excel文件。
    :param:file_path: (str): Excel文件路径
    :param:rows: (list of list): 要追加的行数据，每行是一个列表
    :param:sheet_name: (str): 要追加数据的表格名称，可选参数，默认为None，表示追加到当前活动表格
    """
    # 加载现有的工作簿
    wb = load_workbook(file_path)

    if sheet_name:
        # 检查工作表是否存在
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
    else:
        # 获取活动工作表
        ws = wb.active

        # 追加行数据
    for row in rows:
        ws.append(row)

    # 保存工作簿
    wb.save(file_path)


if __name__ == '__main__':
    file_path = 'excelPath/demo1.xlsx'
    rows = [
        ['申鹤', 40, '成都'],
        ['甘雨', 28, '哈尔滨']
    ]

    append_rows_to_excel(file_path, rows)

    print(f"数据已追加到 {file_path}")
